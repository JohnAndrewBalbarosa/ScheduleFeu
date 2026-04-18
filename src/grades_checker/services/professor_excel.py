from __future__ import annotations

from pathlib import Path
import re

from openpyxl import load_workbook

from grades_checker.models.entities import CourseOfferingSection
from grades_checker.models.entities import SubjectSections

COURSE_CODE_PATTERN = re.compile(r"\b([A-Z]{2,}\d{2,}[A-Z]?|NSTP\d?|OJT|THS\d?)\b")
SECTION_PATTERN = re.compile(r"\b([A-Z]{1,4}\d{2,4}[A-Z]?)\b")

COURSE_HEADER_HINTS = ("course", "subject", "code")
SECTION_HEADER_HINTS = ("section", "sec")
PROF_HEADER_HINTS = ("prof", "instructor", "faculty", "teacher")


def _normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_header_mapping(rows: list[list[str]]) -> tuple[int, int | None, int | None, int | None] | None:
    for row_index, row in enumerate(rows[:30]):
        lowered = [cell.lower() for cell in row]

        course_col = next(
            (idx for idx, cell in enumerate(lowered) if any(hint in cell for hint in COURSE_HEADER_HINTS)),
            None,
        )
        section_col = next(
            (idx for idx, cell in enumerate(lowered) if any(hint in cell for hint in SECTION_HEADER_HINTS)),
            None,
        )
        prof_col = next(
            (idx for idx, cell in enumerate(lowered) if any(hint in cell for hint in PROF_HEADER_HINTS)),
            None,
        )

        if course_col is not None and prof_col is not None:
            return row_index, course_col, section_col, prof_col

        # If explicit course column is not labeled but section + prof are clear,
        # course can still be parsed by regex from row values.
        if section_col is not None and prof_col is not None:
            return row_index, None, section_col, prof_col

    return None


def _extract_course_code(row: list[str], preferred_col: int | None) -> str:
    if preferred_col is not None and preferred_col < len(row):
        candidate = row[preferred_col].upper()
        match = COURSE_CODE_PATTERN.search(candidate)
        if match:
            return match.group(1).upper()

    text = " | ".join(row).upper()
    match = COURSE_CODE_PATTERN.search(text)
    return match.group(1).upper() if match else ""


def _extract_section_code(row: list[str], preferred_col: int | None) -> str:
    if preferred_col is not None and preferred_col < len(row):
        return row[preferred_col].upper()

    text = " | ".join(row).upper()
    match = SECTION_PATTERN.search(text)
    return match.group(1).upper() if match else ""


def _extract_prof_name(row: list[str], prof_col: int | None) -> str:
    if prof_col is not None and prof_col < len(row):
        return row[prof_col].strip()

    # Fallback: pick the longest mostly alphabetic cell.
    best = ""
    for cell in row:
        stripped = cell.strip()
        if len(stripped) < 4:
            continue
        alpha_count = sum(ch.isalpha() for ch in stripped)
        if alpha_count >= max(3, len(stripped) // 2) and len(stripped) > len(best):
            best = stripped
    return best


def load_professor_map_from_excel(xlsx_path: str) -> dict[tuple[str, str], str]:
    workbook = load_workbook(xlsx_path, data_only=True)
    section_map: dict[tuple[str, str], str] = {}
    course_fallback: dict[str, str] = {}

    for sheet in workbook.worksheets:
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([_normalize(cell) for cell in row])

        header = _find_header_mapping(rows)
        if header is None:
            continue

        header_index, course_col, section_col, prof_col = header

        for row in rows[header_index + 1 :]:
            if not any(row):
                continue

            course_code = _extract_course_code(row, course_col)
            if not course_code:
                continue

            section_code = _extract_section_code(row, section_col)
            prof_name = _extract_prof_name(row, prof_col)
            if not prof_name:
                continue

            if section_code:
                section_map[(course_code, section_code)] = prof_name
            course_fallback.setdefault(course_code, prof_name)

    # Populate empty-section fallback keys for convenience.
    for course_code, prof_name in course_fallback.items():
        section_map.setdefault((course_code, ""), prof_name)

    return section_map


def apply_professor_map(
    grouped: list[SubjectSections],
    professor_map: dict[tuple[str, str], str],
) -> list[SubjectSections]:
    updated: list[SubjectSections] = []

    for subject in grouped:
        mapped_sections: list[CourseOfferingSection] = []
        for section in subject.sections:
            exact = professor_map.get((section.course_code.upper(), section.section_code.upper()))
            fallback = professor_map.get((section.course_code.upper(), ""))
            prof_name = exact or fallback or ""

            mapped_sections.append(
                CourseOfferingSection(
                    course_code=section.course_code,
                    section_code=section.section_code,
                    available_slots=section.available_slots,
                    details=section.details,
                    class_size=section.class_size,
                    day=section.day,
                    time=section.time,
                    room=section.room,
                    professor_name=prof_name,
                )
            )

        updated.append(SubjectSections(subject_code=subject.subject_code, sections=mapped_sections))

    return updated


def validate_prof_excel_path(path_text: str | None) -> str | None:
    if not path_text:
        return None
    path = Path(path_text)
    if not path.exists():
        raise FileNotFoundError(f"Professor Excel file not found: {path_text}")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Professor list must be an .xlsx or .xlsm file")
    return str(path)
